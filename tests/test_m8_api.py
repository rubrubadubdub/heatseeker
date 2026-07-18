"""M8 over the UI and JSON API: offerings, lead queue, suppression, XLSX download."""

import io

from fastapi.testclient import TestClient
from heatseeker_api.main import create_app
from heatseeker_common.db import session_scope
from heatseeker_entity_resolution import entities
from heatseeker_entity_resolution.models import ContactType
from heatseeker_intelligence import classifications
from openpyxl import load_workbook


def _seed(engine):
    with session_scope(engine) as session:
        org = entities.create_organisation(
            session, "Acme Scaffolding Pty Ltd", identifiers=[("abn", "51824753556")]
        )
        entities.add_contact_point(session, org, ContactType.ROLE_EMAIL, "info@acme.test")
        classifications.assign(
            session,
            org.id,
            pack_id="scaffolding_anz",
            taxonomy_id="company_archetypes",
            category_id="scaffold_contractor",
            category_label="Scaffold contractor",
            assignment_type="observed",
            confidence=0.8,
        )
        return org.id


def test_offering_to_queue_to_export_roundtrip(engine, settings):
    org_id = _seed(engine)
    client = TestClient(create_app(settings))

    created = client.post(
        "/api/offerings",
        json={
            "name": "Scaffold design outsourcing",
            "target_archetype_ids": ["scaffold_contractor"],
            "need_gap_capability_ids": ["scaffold_design"],
        },
    )
    assert created.status_code == 201
    offering_id = created.json()["id"]

    summary = client.post(f"/api/offerings/{offering_id}/rescore").json()
    assert summary["scored"] == 1

    leads = client.get("/api/leads", params={"offering_id": offering_id}).json()["leads"]
    assert len(leads) == 1
    lead = leads[0]
    assert lead["organisation_id"] == org_id
    assert lead["commercial_priority"] > 0
    assert lead["reasons"] and all("dimension" in r for r in lead["reasons"])
    assert any("M7" in u for u in lead["unknowns"])  # honest timing stub
    assert lead["component_scores"]["weights"]

    exported = client.get("/api/leads/export.xlsx", params={"offering_id": offering_id})
    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    workbook = load_workbook(io.BytesIO(exported.content))
    assert {"Leads", "Method"} == set(workbook.sheetnames)

    page = client.get(f"/leads?offering_id={offering_id}")
    assert page.status_code == 200
    assert "Acme Scaffolding Pty Ltd" in page.text
    assert "Export XLSX" in page.text

    entity_page = client.get(f"/entities/{org_id}")
    assert "Commercial action" in entity_page.text
    assert "Outreach is never automatic" in entity_page.text


def test_ui_offering_form_suppress_and_lift(engine, settings):
    org_id = _seed(engine)
    client = TestClient(create_app(settings))

    response = client.post(
        "/leads/offerings",
        data={
            "name": "Drafting outsourcing",
            "target_archetypes": "scaffold_contractor",
            "geo_codes": "AU",
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    offering_id = client.get("/api/offerings").json()["offerings"][0]["id"]
    client.post(f"/api/offerings/{offering_id}/rescore")

    suppressed = client.post(
        "/leads/suppress",
        data={"organisation_id": org_id, "offering_id": offering_id, "reason": "opt_out"},
        follow_redirects=False,
    )
    assert suppressed.status_code == 303
    queue = client.get("/api/leads", params={"offering_id": offering_id}).json()["leads"]
    assert queue == []  # suppression respected in the queue

    exported = client.get("/api/leads/export.xlsx", params={"offering_id": offering_id})
    workbook = load_workbook(io.BytesIO(exported.content))
    names = [row[3].value for row in workbook["Leads"].iter_rows(min_row=2)]
    assert "Acme Scaffolding Pty Ltd" not in names  # and in the export

    rule = client.post(
        f"/api/organisations/{org_id}/suppress", json={"reason": "opt_out"}
    ).json()
    lifted = client.post(f"/api/suppressions/{rule['rule_id']}/lift")
    assert lifted.status_code == 200
    client.post(f"/api/offerings/{offering_id}/rescore")
    queue = client.get("/api/leads", params={"offering_id": offering_id}).json()["leads"]
    assert len(queue) == 1

    dashboard = client.get("/")
    assert "lead queue" in dashboard.text.lower()


def test_api_guards(engine, settings):
    client = TestClient(create_app(settings))
    assert client.post("/api/offerings", json={"name": "  "}).status_code == 422
    assert client.post("/api/offerings/nope/rescore").status_code == 404
    assert client.get("/api/leads", params={"offering_id": "nope"}).status_code == 404
    assert (
        client.get("/api/leads/export.xlsx", params={"offering_id": "nope"}).status_code
        == 404
    )
