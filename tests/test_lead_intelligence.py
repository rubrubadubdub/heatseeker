"""Lead engine: explained scoring, suppression, XLSX export (M8, spec §19/§20/§32.3)."""

import io

import pytest
from heatseeker_common.db import session_scope
from heatseeker_common.timeutil import utc_now
from heatseeker_entity_resolution import entities
from heatseeker_entity_resolution.models import ContactType
from heatseeker_intelligence import capabilities, classifications, facts
from heatseeker_intelligence.observations import record_observation
from heatseeker_lead_intelligence import service
from heatseeker_lead_intelligence.export import build_lead_workbook
from heatseeker_lead_intelligence.models import OpportunityStage
from heatseeker_lead_intelligence.scoring import score_organisation
from openpyxl import load_workbook
from test_intelligence_facts import make_document, make_source


def _offering(session, **overrides):
    defaults = dict(
        name="Scaffold design & drafting outsourcing",
        target_archetype_ids=["scaffold_contractor"],
        target_capability_ids=["scaffold_erection"],
        need_gap_capability_ids=["scaffold_design"],
        negative_archetype_ids=["temporary_works_designer"],
        geo_codes=["AU-QLD"],
    )
    defaults.update(overrides)
    return service.create_offering(session, **defaults)


def _classified_org(session, name, *, archetype="scaffold_contractor", locality="Brisbane"):
    org = entities.create_organisation(session, name)
    location = entities.add_location(
        session, locality=locality, region="QLD", country="AU", location_type="office"
    )
    entities.set_primary_location(session, org, location)
    classifications.assign(
        session,
        org.id,
        pack_id="scaffolding_anz",
        taxonomy_id="company_archetypes",
        category_id=archetype,
        category_label=archetype,
        assignment_type="observed",
        confidence=0.8,
    )
    return org


def _add_capability(session, org, capability_id, source, *, contradicts=False):
    document = make_document(session, source, f"{org.id}-{capability_id}")
    observation = record_observation(
        session, document, "service_claim", capability_id, subject_entity_id=org.id
    )
    return capabilities.record_capability_evidence(
        session,
        org.id,
        pack_id="scaffolding_anz",
        capability_id=capability_id,
        capability_label=capability_id,
        observation_id=observation.id,
        source_definition_id=source.id,
        source_category=source.source_category,
        authority_tier=source.authority_tier,
        observed_at=utc_now(),
        contradicts=contradicts,
    )


def test_scores_are_explained_with_evidence_and_unknowns(engine):
    with session_scope(engine) as session:
        offering = _offering(session)
        org = _classified_org(session, "Acme Scaffolding")
        project_source = make_source(session, "Project page", category="project_registry")
        _add_capability(session, org, "scaffold_erection", project_source)
        entities.add_contact_point(
            session, org, ContactType.ROLE_EMAIL, "estimating@acme.com.au"
        )

        score = score_organisation(session, org, offering)
        dimensions = {reason["dimension"] for reason in score.reasons}
        assert {"industry_fit", "service_fit", "geographic_fit", "accessibility"} <= dimensions
        # Evidence references attached, not just prose (§19.6).
        industry = next(r for r in score.reasons if r["dimension"] == "industry_fit")
        assert industry["evidence"]
        # Need-gap absence is a hypothesis + an unknown, never a confirmed fact (§6.3).
        assert any("no visible internal capability" in r["text"] for r in score.reasons)
        assert any("no reliable evidence of internal staffing" in u for u in score.unknowns)
        # Timing is a declared stub until M7 (ADR-0015).
        assert score.timing == 0.5
        assert any("M7" in u for u in score.unknowns)
        assert 0 < score.commercial_priority <= 1
        assert score.components["weights"]  # formula inputs stored, inspectable


def test_weak_evidence_lowers_priority(engine):
    with session_scope(engine) as session:
        offering = _offering(session)
        strong = _classified_org(session, "Strong Evidence Co")
        weak = _classified_org(session, "Weak Evidence Co")
        for org in (strong, weak):
            entities.add_contact_point(
                session, org, ContactType.ROLE_EMAIL, f"info@{org.id[:6]}.example"
            )
        site = make_source(session, "Strong site")
        registry = make_source(session, "Registry", category="government_registry", tier=1)
        for source in (site, registry):
            record_observation(
                session,
                make_document(session, source, strong.id),
                "phone",
                "+61 7 3333 1111",
                subject_entity_id=strong.id,
            )
        facts.reconcile(session, strong.id, "phone")

        strong_score = score_organisation(session, strong, offering)
        weak_score = score_organisation(session, weak, offering)
        assert strong_score.evidence_quality > weak_score.evidence_quality
        assert strong_score.commercial_priority > weak_score.commercial_priority
        assert any("no reconciled facts" in u for u in weak_score.unknowns)


def test_negative_indicators_and_gap_presence(engine):
    with session_scope(engine) as session:
        offering = _offering(session)
        competitor = _classified_org(
            session, "Design Rival", archetype="temporary_works_designer"
        )
        score = score_organisation(session, competitor, offering)
        assert any("negative archetype" in risk for risk in score.risks)

        equipped = _classified_org(session, "Self-Sufficient Scaffolds")
        project_source = make_source(session, "Projects", category="project_registry")
        _add_capability(session, equipped, "scaffold_design", project_source)
        equipped_score = score_organisation(session, equipped, offering)
        # In-house design = secondary opportunity (overflow/supplant), not disqualified.
        assert equipped_score.components["need_likelihood"] == 0.45
        assert any("secondary target" in r["text"] for r in equipped_score.reasons)

        merged = entities.create_organisation(session, "Ghost Co")
        merged.status = "inactive"
        assert score_organisation(session, merged, offering).skip is True


def test_rescore_suppression_and_lift(engine):
    with session_scope(engine) as session:
        offering = _offering(session)
        org = _classified_org(session, "Acme Scaffolding")
        entities.add_contact_point(session, org, ContactType.GENERAL_EMAIL, "hi@acme.test")
        offering_id, org_id = offering.id, org.id

    with session_scope(engine) as session:
        summary = service.rescore_offering(session, offering_id)
        assert summary["scored"] == 1
        queue = service.lead_queue(session, offering_id)
        assert len(queue) == 1
        assert queue[0].commercial_priority > 0

    with session_scope(engine) as session:
        rule = service.suppress(session, org_id, reason="do_not_contact", note="asked us")
        rule_id = rule.id
        # Existing lead collapses immediately; a rescore keeps it suppressed (§32.3).
        service.rescore_offering(session, offering_id)
        assert service.lead_queue(session, offering_id) == []
        suppressed = service.lead_queue(session, offering_id, include_suppressed=True)
        assert suppressed[0].opportunity_stage == OpportunityStage.SUPPRESSED
        assert suppressed[0].commercial_priority == 0.0
        assert any("suppressed: do_not_contact" in r for r in suppressed[0].risks)

    with session_scope(engine) as session:
        service.lift_suppression(session, rule_id)
        service.rescore_offering(session, offering_id)
        queue = service.lead_queue(session, offering_id)
        assert len(queue) == 1
        assert queue[0].commercial_priority > 0


def test_xlsx_export_full_columns_and_suppression(engine):
    with session_scope(engine) as session:
        offering = _offering(session)
        lead_org = _classified_org(session, "Acme Scaffolding Pty Ltd")
        lead_org.description = "Commercial scaffolding contractor"
        entities.add_identifier(session, lead_org, "abn", "51824753556")
        entities.add_domain(session, lead_org, "acme.com.au")
        entities.add_contact_point(
            session, lead_org, ContactType.ROLE_EMAIL, "estimating@acme.com.au"
        )
        hidden = _classified_org(session, "Hidden Co")
        service.suppress(session, hidden.id, reason="opt_out")
        service.rescore_offering(session, offering.id)

        payload = build_lead_workbook(session, offering)

    workbook = load_workbook(io.BytesIO(payload))
    assert set(workbook.sheetnames) == {"Leads", "Method"}
    sheet = workbook["Leads"]
    headers = [cell.value for cell in sheet[1]]
    for expected in (
        "commercial_priority", "canonical_name", "identifiers", "domains",
        "best_contact_route", "reasons", "risks", "unknowns", "capabilities",
        "operating_tier", "profile_url", "rule_version",
    ):
        assert expected in headers
    names = [row[headers.index("canonical_name")].value for row in sheet.iter_rows(min_row=2)]
    assert "Acme Scaffolding Pty Ltd" in names
    assert "Hidden Co" not in names  # suppression respected in exports (§32.3)

    row = next(
        row for row in sheet.iter_rows(min_row=2)
        if row[headers.index("canonical_name")].value == "Acme Scaffolding Pty Ltd"
    )
    assert row[headers.index("identifiers")].value == "abn:51824753556"
    assert row[headers.index("best_contact_route")].value == "role_email"
    assert row[headers.index("unknowns")].value  # timing stub etc. carried into the file

    method = workbook["Method"]
    method_text = "\n".join(str(row[1].value) for row in method.iter_rows())
    assert "Suppressed organisations are excluded" in method_text
    assert "human decision" in method_text


def test_offering_validation(engine):
    with session_scope(engine) as session, pytest.raises(ValueError):
        service.create_offering(session, "   ")
