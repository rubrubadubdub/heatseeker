"""XLSX export of one offering's lead queue (pulled forward from M10 — ADR-0015).

Suppressed organisations are absent from the workbook entirely (§32.3), and a Method
sheet states the scoring formula, rule version, offering configuration, and generation
time so the file is self-describing evidence, not a bare list.
"""

import io
from datetime import datetime

from heatseeker_common.timeutil import utc_now
from heatseeker_entity_resolution.resolution import merge_group
from heatseeker_intelligence.capabilities import capabilities_for
from heatseeker_intelligence.classifications import classifications_for
from heatseeker_intelligence.models import SizeConcept
from heatseeker_intelligence.sizing import estimates_for
from heatseeker_knowledge_graph.graph import edges_for
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from heatseeker_lead_intelligence.models import Offering
from heatseeker_lead_intelligence.scoring import PRIORITY_MIX, SCORING_RULE_VERSION
from heatseeker_lead_intelligence.service import lead_queue

# §20.2 order for the contact columns.
_CONTACT_ORDER = (
    "role_email",
    "general_email",
    "contact_form",
    "enquiry_url",
    "phone",
    "social_profile",
    "postal_address",
    "other",
)

HEADERS = (
    "rank",
    "commercial_priority",
    "stage",
    "canonical_name",
    "legal_name",
    "trading_names",
    "operating_tier",
    "legal_entity_size",
    "operating_group_size",
    "identifiers",
    "domains",
    "locality",
    "region",
    "country",
    "best_contact_route",
    "contacts_public",
    "fit_score",
    "industry_fit",
    "service_fit",
    "need_likelihood",
    "scale",
    "geographic_fit",
    "timing_score",
    "evidence_quality",
    "accessibility",
    "reasons",
    "risks",
    "unknowns",
    "next_action",
    "capabilities",
    "classifications",
    "connection_count",
    "profile_completeness",
    "entity_confidence",
    "profile_url",
    "scored_at",
    "rule_version",
)


def _fmt_dt(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%d %H:%M") if value else ""


def _join(parts) -> str:
    return "; ".join(str(p) for p in parts if p)


def build_lead_workbook(session: Session, offering: Offering) -> bytes:
    """Everything Heatseeker knows about each non-suppressed lead, one row each."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    sheet.append(list(HEADERS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    leads = lead_queue(session, offering.id, include_suppressed=False)
    for rank, lead in enumerate(leads, start=1):
        organisation = lead.organisation
        group = merge_group(session, organisation.id)
        group_ids = [o.id for o in group]
        components = lead.component_scores or {}
        sizes = {e.concept: e.band for e in estimates_for(session, group_ids)}
        capabilities = capabilities_for(session, group_ids)
        classifications = classifications_for(session, group_ids)
        location = organisation.primary_location

        contacts = [
            c for o in group for c in o.contact_points if c.public_business_contact
        ]
        contacts.sort(
            key=lambda c: _CONTACT_ORDER.index(c.contact_type)
            if c.contact_type in _CONTACT_ORDER
            else len(_CONTACT_ORDER)
        )
        best_route = contacts[0].contact_type if contacts else ""

        sheet.append(
            [
                rank,
                lead.commercial_priority,
                lead.opportunity_stage,
                organisation.canonical_name,
                organisation.legal_name or "",
                _join(organisation.trading_names or []),
                sizes.get(SizeConcept.CAPABILITY_TIER, "unresolved"),
                sizes.get(SizeConcept.LEGAL_ENTITY_SIZE, "unresolved"),
                sizes.get(SizeConcept.OPERATING_GROUP_SIZE, "unresolved"),
                _join(f"{i.scheme}:{i.value}" for o in group for i in o.identifiers),
                _join(d.domain for o in group for d in o.domains),
                (location.locality if location else "") or "",
                (location.region if location else "") or "",
                (location.country if location else "") or "",
                best_route,
                _join(f"{c.contact_type}:{c.value}" for c in contacts),
                lead.fit_score,
                components.get("industry_fit", ""),
                components.get("service_fit", ""),
                components.get("need_likelihood", ""),
                components.get("scale", ""),
                components.get("geographic_fit", ""),
                lead.timing_score,
                lead.evidence_quality_score,
                lead.accessibility_score,
                _join(f"[{r['dimension']}] {r['text']}" for r in lead.reasons),
                _join(lead.risks),
                _join(lead.unknowns),
                (lead.next_action or {}).get("text", ""),
                _join(
                    f"{c.capability_label or c.capability_id} ({c.capability_status})"
                    for c in capabilities
                ),
                _join(
                    f"{c.taxonomy_id}:{c.category_label or c.category_id} "
                    f"({c.assignment_type})"
                    for c in classifications
                ),
                len(edges_for(session, organisation.id)),
                organisation.profile_completeness,
                organisation.entity_confidence,
                f"/entities/{organisation.id}",
                _fmt_dt(lead.scored_at),
                lead.rule_version,
            ]
        )

    for index, header in enumerate(HEADERS, start=1):
        width = 14
        if header in ("canonical_name", "legal_name", "reasons", "risks", "unknowns",
                      "capabilities", "classifications", "contacts_public"):
            width = 40
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"

    method = workbook.create_sheet("Method")
    method_rows = [
        ("Generated", _fmt_dt(utc_now())),
        ("Offering", offering.name),
        ("Offering description", offering.description or ""),
        ("Rule version", SCORING_RULE_VERSION),
        ("Leads exported", len(leads)),
        (
            "Priority formula",
            "commercial_priority = (fit·{fit} + timing·{timing} + evidence·{evidence} "
            "+ accessibility·{accessibility}) x (0.5 + 0.5·evidence_quality)".format(
                **PRIORITY_MIX
            ),
        ),
        ("Timing note", "Timing is a neutral stub until M7 news/events land (ADR-0015)."),
        (
            "Suppression",
            "Suppressed organisations are excluded from this file entirely (§32.3).",
        ),
        (
            "Provenance",
            "Every score derives from evidence-backed records; open a profile_url for "
            "field-level evidence and confidence.",
        ),
        ("No automatic outreach", "Contacting anyone is a human decision (§19.7)."),
        ("Target archetypes", _join(offering.target_archetype_ids)),
        ("Target capabilities", _join(offering.target_capability_ids)),
        ("Need-gap capabilities", _join(offering.need_gap_capability_ids)),
        ("Negative archetypes", _join(offering.negative_archetype_ids)),
        ("Geography", _join(offering.geo_codes) or "active scope / unrestricted"),
    ]
    for row in method_rows:
        method.append(list(row))
    method.column_dimensions["A"].width = 26
    method.column_dimensions["B"].width = 110
    for cell in method["A"]:
        cell.font = Font(bold=True)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
